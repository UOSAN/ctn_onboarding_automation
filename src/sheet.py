from typing import Union

from openpyxl import load_workbook

from src.person import Person

NUM_COLUMNS = 45


def string_equal_ish(a: str, b: Union[str, None]):
    return a == b or (len(a) == 0 and b is None)


class Sheet:
    def __init__(self, file_name: str):
        self._wb = load_workbook(filename=file_name)
        self._file_name = file_name

    def add_person(self, person: Person, note: str):
        ws = self._wb.active
        ws.append(person.to_list(note))
        self._wb.save(self._file_name)

    def find_person(self, person: Person):
        ws = self._wb.active
        found = False
        for row in ws.iter_rows(max_col=2):
            last_name, first_name = row
            if string_equal_ish(person.last_name, last_name.value)\
                    and string_equal_ish(person.first_name, first_name.value):
                found = True
                break

        return found

    def size_changed(self):
        ws = self._wb.active
        return ws.max_column != NUM_COLUMNS
